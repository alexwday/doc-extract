"""Excel export service - generates Excel files from extraction results."""

import io
import uuid
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .storage import storage
from ..config import EXPORTS_DIR


# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelService:
    """Service for generating Excel exports from extraction results."""

    def generate_export(
        self,
        result_ids: list[str],
        filename: Optional[str] = None,
    ) -> dict:
        """
        Generate an Excel export from extraction results.

        Args:
            result_ids: List of result IDs to include
            filename: Optional custom filename (without .xlsx)

        Returns:
            Export metadata dict
        """
        # Load all results
        results = []
        for result_id in result_ids:
            result = storage.get_result(result_id)
            if result:
                results.append(result)

        if not results:
            raise ValueError("No valid results found")

        # Create workbook
        wb = Workbook()

        # Create sheets
        self._create_extractions_sheet(wb, results)
        self._create_summaries_sheet(wb, results)
        self._create_entity_group_sheets(wb, results)
        self._create_metadata_sheet(wb, results)

        # Remove default sheet if it exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Generate export ID and filename
        export_id = str(uuid.uuid4())
        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"extraction_export_{timestamp}"

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.getvalue()

        # Save to file
        export_path = EXPORTS_DIR / f"{export_id}.xlsx"
        with open(export_path, "wb") as f:
            f.write(content)

        return {
            "id": export_id,
            "filename": f"{filename}.xlsx",
            "created_at": datetime.utcnow().isoformat() + "Z",
            "result_count": len(results),
            "file_size_bytes": len(content),
            "path": str(export_path),
        }

    def _create_extractions_sheet(self, wb: Workbook, results: list[dict]):
        """Create the Extractions sheet with metric values and bounding boxes."""
        ws = wb.create_sheet("Extractions", 0)

        # Headers
        headers = [
            "Document",
            "Document ID",
            "Field",
            "Display Name",
            "Value",
            "Page",
            "Confidence",
            "Verified",
            "X1",
            "Y1",
            "X2",
            "Y2",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Data rows
        row = 2
        for result in results:
            doc_name = result.get("document_name", "")
            doc_id = result.get("document_id", "")

            for field_name, extraction in result.get("extractions", {}).items():
                # Skip summaries (they go in the other sheet)
                if extraction.get("field_type") == "summary":
                    continue

                bbox = extraction.get("bounding_box") or {}

                ws.cell(row=row, column=1, value=doc_name)
                ws.cell(row=row, column=2, value=doc_id)
                ws.cell(row=row, column=3, value=field_name)
                ws.cell(row=row, column=4, value=extraction.get("display_name", ""))
                ws.cell(row=row, column=5, value=extraction.get("value", ""))
                ws.cell(row=row, column=6, value=extraction.get("page", ""))
                ws.cell(row=row, column=7, value=extraction.get("confidence", 0))
                ws.cell(row=row, column=8, value="Yes" if extraction.get("verified") else "No")
                ws.cell(row=row, column=9, value=bbox.get("x1", ""))
                ws.cell(row=row, column=10, value=bbox.get("y1", ""))
                ws.cell(row=row, column=11, value=bbox.get("x2", ""))
                ws.cell(row=row, column=12, value=bbox.get("y2", ""))

                # Apply borders
                for col in range(1, 13):
                    ws.cell(row=row, column=col).border = THIN_BORDER

                row += 1

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_summaries_sheet(self, wb: Workbook, results: list[dict]):
        """Create the Summaries sheet with synthesized content."""
        ws = wb.create_sheet("Summaries", 1)

        # Headers
        headers = [
            "Document",
            "Document ID",
            "Field",
            "Display Name",
            "Summary",
            "Source Pages",
            "Confidence",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Data rows
        row = 2
        for result in results:
            doc_name = result.get("document_name", "")
            doc_id = result.get("document_id", "")

            for field_name, extraction in result.get("extractions", {}).items():
                # Only include summaries
                if extraction.get("field_type") != "summary":
                    continue

                source_pages = extraction.get("source_pages", [])
                source_pages_str = ", ".join(str(p) for p in source_pages) if source_pages else ""

                ws.cell(row=row, column=1, value=doc_name)
                ws.cell(row=row, column=2, value=doc_id)
                ws.cell(row=row, column=3, value=field_name)
                ws.cell(row=row, column=4, value=extraction.get("display_name", ""))

                # Summary cell with text wrapping
                summary_cell = ws.cell(row=row, column=5, value=extraction.get("value", ""))
                summary_cell.alignment = Alignment(wrap_text=True, vertical="top")

                ws.cell(row=row, column=6, value=source_pages_str)
                ws.cell(row=row, column=7, value=extraction.get("confidence", 0))

                # Apply borders
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = THIN_BORDER

                row += 1

        # Auto-adjust column widths (but cap summary column)
        self._auto_adjust_columns(ws, max_width=80)

        # Set summary column to a reasonable width
        ws.column_dimensions["E"].width = 80

    def _create_entity_group_sheets(self, wb: Workbook, results: list[dict]):
        """Create sheets for entity groups - one sheet per unique group across all results."""
        # Collect all unique entity groups across all results
        all_groups = {}  # group_name -> {display_name, field_names}
        for result in results:
            for group_name, group_data in result.get("entity_groups", {}).items():
                if group_name not in all_groups:
                    # Get field names from first entity in the group
                    field_names = []
                    if group_data.get("entities"):
                        field_names = list(group_data["entities"][0].get("values", {}).keys())
                    all_groups[group_name] = {
                        "display_name": group_data.get("display_name", group_name),
                        "field_names": field_names,
                    }

        # Create a sheet for each entity group
        sheet_index = 2  # After Extractions and Summaries
        for group_name, group_info in all_groups.items():
            sheet_name = group_info["display_name"][:31]  # Excel sheet name limit
            ws = wb.create_sheet(sheet_name, sheet_index)
            sheet_index += 1

            # Headers: Document, Entity #, Page, Confidence, [field columns]
            headers = ["Document", "Document ID", "Entity #", "Page", "Confidence"]
            headers.extend(group_info["field_names"])

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER

            # Data rows
            row = 2
            for result in results:
                doc_name = result.get("document_name", "")
                doc_id = result.get("document_id", "")
                group_data = result.get("entity_groups", {}).get(group_name, {})

                for entity_idx, entity in enumerate(group_data.get("entities", []), 1):
                    ws.cell(row=row, column=1, value=doc_name)
                    ws.cell(row=row, column=2, value=doc_id)
                    ws.cell(row=row, column=3, value=entity_idx)
                    ws.cell(row=row, column=4, value=entity.get("page", ""))
                    ws.cell(row=row, column=5, value=entity.get("confidence", 0))

                    # Add field values
                    values = entity.get("values", {})
                    for col_offset, field_name in enumerate(group_info["field_names"]):
                        ws.cell(row=row, column=6 + col_offset, value=values.get(field_name, ""))

                    # Apply borders
                    for col in range(1, 6 + len(group_info["field_names"])):
                        ws.cell(row=row, column=col).border = THIN_BORDER

                    row += 1

            # Auto-adjust column widths
            self._auto_adjust_columns(ws)

    def _create_metadata_sheet(self, wb: Workbook, results: list[dict]):
        """Create the Document Info sheet with metadata."""
        ws = wb.create_sheet("Document Info", 2)

        # Headers
        headers = [
            "Document",
            "Document ID",
            "Template",
            "Template ID",
            "Pages",
            "Extracted At",
            "Processing Time (s)",
            "Verified",
            "Extraction Count",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Data rows
        row = 2
        for result in results:
            # Get document info for page count
            doc = storage.get_document(result.get("document_id", ""))
            page_count = doc.get("page_count", "") if doc else ""

            ws.cell(row=row, column=1, value=result.get("document_name", ""))
            ws.cell(row=row, column=2, value=result.get("document_id", ""))
            ws.cell(row=row, column=3, value=result.get("template_name", ""))
            ws.cell(row=row, column=4, value=result.get("template_id", ""))
            ws.cell(row=row, column=5, value=page_count)
            ws.cell(row=row, column=6, value=result.get("extracted_at", ""))
            ws.cell(row=row, column=7, value=result.get("processing_time_seconds", 0))
            ws.cell(row=row, column=8, value="Yes" if result.get("verified") else "No")
            ws.cell(row=row, column=9, value=len(result.get("extractions", {})))

            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = THIN_BORDER

            row += 1

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _auto_adjust_columns(self, ws, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value or "")) for cell in column_cells
            )
            # Add some padding and cap at max_width
            adjusted_width = min(length + 2, max_width)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    def get_export(self, export_id: str) -> Optional[tuple[bytes, str]]:
        """
        Get an export file by ID.

        Args:
            export_id: Export ID

        Returns:
            Tuple of (file_content, filename) or None if not found
        """
        export_path = storage.get_export_path(export_id)
        if not export_path:
            return None

        with open(export_path, "rb") as f:
            content = f.read()

        return content, f"{export_id}.xlsx"


# Singleton instance
excel_service = ExcelService()
